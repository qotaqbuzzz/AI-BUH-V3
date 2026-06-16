#!/usr/bin/env python3
"""
Python rendering sidecar for agent-deepseek.

Reads JSON payload from stdin, writes JSON result to stdout.
Usage:  python render_cli.py <mode>
Modes:  segments | xlsx | pdf

segments  → { segments: [{type:"text",md:...} | {type:"png",b64:...}] }
xlsx      → { b64: "<base64 workbook>" }
pdf       → { b64: "<base64 pdf>" }
"""
from __future__ import annotations

import sys
import json
import base64
import re
import textwrap
import html as _html_mod
from io import BytesIO

# ── Matplotlib (headless) ─────────────────────────────────────────────────────
import matplotlib
matplotlib.use("Agg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.patches as mpatches

# ══════════════════════════════════════════════════════════════════════════════
# Colour palette  (mirrors accountant_bot/bot/table_render.py)
# ══════════════════════════════════════════════════════════════════════════════
_C = {
    "canvas":       "#F8F9FB",
    "header_bg":    "#1E2D45",
    "header_fg":    "#FFFFFF",
    "title_bg":     "#2C3E50",
    "title_fg":     "#FFFFFF",
    "row_odd":      "#FFFFFF",
    "row_even":     "#F2F5FA",
    "total_bg":     "#E4EBF5",
    "total_border": "#1E2D45",
    "text":         "#1A2332",
    "muted":        "#9BA8B5",
    "negative":     "#C0392B",
    "account":      "#2563A8",
    "grid":         "#DDE3EC",
}

# ══════════════════════════════════════════════════════════════════════════════
# Regex helpers  (ported verbatim from accountant_bot/bot/table_render.py)
# ══════════════════════════════════════════════════════════════════════════════
_TOTAL_KW = re.compile(
    r"\b(итого|всего|баланс|total|subtotal|прибыль|убыток|profit|loss|net)\b",
    re.IGNORECASE,
)
_NUM_RE   = re.compile(r"^[-–+(]?[\d\s,.' ₸₽$€%()–\-]+[₸₽]?\)?$")
_NEG_RE   = re.compile(r"^\s*[\(\-–]")
_ZERO_RE  = re.compile(r"^[\s0,. ₸₽]+$")
_ACCT_RE  = re.compile(r"^\d{3,5}$")
_SEP_RE   = re.compile(r"[|:\-\s]")
_MD_RE    = re.compile(r"(\*{1,3}|_{1,3}|~~|`|#+\s*)")
_EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001FAFF"
    "\U00002600-\U000027BF"
    "⌀-⏿"
    "⬀-⯿"
    "■-◿"
    "]+",
    re.UNICODE,
)

# ══════════════════════════════════════════════════════════════════════════════
# Markdown-table parsing helpers
# ══════════════════════════════════════════════════════════════════════════════

def _is_sep(line: str) -> bool:
    s = line.strip()
    return "|" in s and "-" in s and _SEP_RE.sub("", s) == ""


def _strip_md(text: str) -> str:
    t = _MD_RE.sub("", text)
    t = _EMOJI_RE.sub("", t)
    return t.strip()


def _parse_cells(line: str) -> list[str]:
    return [_strip_md(c) for c in line.strip().strip("|").split("|")]


def _strip_emoji(s: str | None) -> str | None:
    if s is None:
        return None
    return _EMOJI_RE.sub("", s).strip() or None


# ══════════════════════════════════════════════════════════════════════════════
# Ordered segment splitter  (new: preserves original text/table order)
# ══════════════════════════════════════════════════════════════════════════════

def split_ordered(text: str, min_cols: int = 2) -> list[dict]:
    """
    Split *text* into an ordered list of segments:
        {"type": "text",  "md": "…"}
        {"type": "table", "rows": [[…], …], "title": str | None}

    Wide tables (≥ min_cols columns) are extracted as "table" segments.
    All other content stays as "text" segments in original order.
    """
    lines = text.splitlines(keepends=True)
    segments: list[dict] = []
    text_buf: list[str] = []
    i = 0

    def flush_text() -> None:
        if text_buf:
            content = "".join(text_buf)
            if content.strip():
                segments.append({"type": "text", "md": content})
            text_buf.clear()

    while i < len(lines):
        line = lines[i]
        # Detect table header row followed by a separator
        if "|" in line and i + 1 < len(lines) and _is_sep(lines[i + 1]):
            cells = _parse_cells(line)
            if len(cells) >= min_cols:
                # Look back in text_buf for nearest non-empty heading as title
                title: str | None = None
                for prev_line in reversed(text_buf):
                    candidate = _strip_md(prev_line)
                    if candidate:
                        title = candidate
                        break

                rows: list[list[str]] = [cells]
                i += 2  # skip header + separator
                while i < len(lines) and "|" in lines[i] and not _is_sep(lines[i]):
                    rows.append(_parse_cells(lines[i]))
                    i += 1

                flush_text()
                segments.append({"type": "table", "rows": rows, "title": title})
                continue

        text_buf.append(line)
        i += 1

    flush_text()
    return segments


# ══════════════════════════════════════════════════════════════════════════════
# Column / row analysis  (ported from accountant_bot/bot/table_render.py)
# ══════════════════════════════════════════════════════════════════════════════

def _detect_col_types(
    data: list[list[str]], n_cols: int
) -> tuple[list[bool], list[bool]]:
    numeric, account = [], []
    for c in range(n_cols):
        vals = [r[c] for r in data if r[c].strip() and r[c].strip() != "—"]
        num_hit = sum(1 for v in vals if _NUM_RE.match(v))
        acc_hit = sum(1 for v in vals if _ACCT_RE.match(v.strip()))
        n = max(len(vals), 1)
        numeric.append(num_hit / n >= 0.5)
        account.append(acc_hit / n >= 0.6)
    return numeric, account


def _is_total_row(row: list[str]) -> bool:
    return any(_TOTAL_KW.search(cell) for cell in row)


def _fmt_cell(val: str, is_numeric: bool) -> tuple[str, str]:
    v = val.strip()
    if not v:
        return ("—", _C["muted"]) if is_numeric else ("", _C["text"])
    if is_numeric and _ZERO_RE.match(v) and v not in ("0", ""):
        return "—", _C["muted"]
    if is_numeric and _NEG_RE.match(v):
        inner = v.lstrip("-(–").rstrip(")")
        return f"({inner})", _C["negative"]
    return v, _C["text"]


# ══════════════════════════════════════════════════════════════════════════════
# Figure builder  (ported verbatim from accountant_bot/bot/table_render.py)
# ══════════════════════════════════════════════════════════════════════════════

def _build_figure(
    rows: list[list[str]],
    dpi: int = 150,
    title: str | None = None,
) -> Figure:
    n_cols = max(len(r) for r in rows)
    norm   = [(r + [""] * n_cols)[:n_cols] for r in rows]
    header    = norm[0]
    data_rows = norm[1:] if len(norm) > 1 else [[""] * n_cols]

    col_numeric, col_account = _detect_col_types(data_rows, n_cols)
    total_set = {i for i, r in enumerate(data_rows) if _is_total_row(r)}

    CHAR_W   = 0.082
    PAD_H    = 0.13
    V_PAD    = 0.08
    LINE_H   = 0.20
    HEAD_H   = 0.32
    TITLE_H  = 0.36 if title else 0.0
    MIN_W    = 0.50
    MAX_TXT  = 3.00
    MAX_NUM  = 1.80
    MARGIN   = 0.18

    col_widths: list[float] = []
    for c in range(n_cols):
        max_chars = max(len(header[c]), max((len(r[c]) for r in data_rows), default=0))
        w = max(max_chars * CHAR_W + 2 * PAD_H, MIN_W)
        w = min(w, MAX_NUM if col_numeric[c] else MAX_TXT)
        col_widths.append(w)

    wrap_chars = [max(int((cw - 2 * PAD_H) / CHAR_W), 6) for cw in col_widths]

    def _wrap(text_: str, col: int) -> list[str]:
        if not text_:
            return [""]
        return textwrap.wrap(text_, width=wrap_chars[col]) or [text_]

    wrapped: list[list[list[str]]] = []
    row_heights: list[float] = []
    for row in data_rows:
        wr = [_wrap(_fmt_cell(row[c], col_numeric[c])[0], c) for c in range(n_cols)]
        wrapped.append(wr)
        n_lines = max(len(lines_) for lines_ in wr)
        row_heights.append(max(n_lines * LINE_H + V_PAD, LINE_H + V_PAD))

    tbl_w = sum(col_widths)
    tbl_h = TITLE_H + HEAD_H + sum(row_heights)
    fig_w = tbl_w + 2 * MARGIN
    fig_h = tbl_h + 2 * MARGIN

    fig = Figure(figsize=(fig_w, fig_h), dpi=dpi, facecolor=_C["canvas"])
    ax  = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.set_axis_off()
    ax.set_facecolor(_C["canvas"])

    y_cursor = fig_h - MARGIN

    if title:
        y_cursor -= TITLE_H
        ax.add_patch(mpatches.FancyBboxPatch(
            (MARGIN, y_cursor), tbl_w, TITLE_H,
            boxstyle="square,pad=0",
            facecolor=_C["title_bg"], edgecolor="none", zorder=1,
        ))
        ax.text(
            MARGIN + tbl_w / 2, y_cursor + TITLE_H / 2,
            title, ha="center", va="center",
            fontsize=10, fontweight="bold",
            color=_C["title_fg"], fontfamily="DejaVu Sans", zorder=2,
        )

    y_cursor -= HEAD_H
    x = MARGIN
    for c, cw in enumerate(col_widths):
        ax.add_patch(mpatches.FancyBboxPatch(
            (x, y_cursor), cw, HEAD_H, boxstyle="square,pad=0",
            facecolor=_C["header_bg"], edgecolor=_C["grid"], linewidth=0.5, zorder=1,
        ))
        align = "right" if col_numeric[c] else "left"
        tx = (x + cw - PAD_H) if col_numeric[c] else (x + PAD_H)
        ax.text(tx, y_cursor + HEAD_H / 2, header[c],
                ha=align, va="center", fontsize=9, fontweight="bold",
                color=_C["header_fg"], fontfamily="DejaVu Sans", zorder=2)
        x += cw

    for i, (row, wr, rh) in enumerate(zip(data_rows, wrapped, row_heights)):
        is_total = i in total_set
        ry = y_cursor - rh
        y_cursor -= rh

        bg = _C["total_bg"] if is_total else (_C["row_odd"] if i % 2 == 0 else _C["row_even"])
        x = MARGIN
        for c, cw in enumerate(col_widths):
            ax.add_patch(mpatches.FancyBboxPatch(
                (x, ry), cw, rh, boxstyle="square,pad=0",
                facecolor=bg, edgecolor=_C["grid"], linewidth=0.3, zorder=1,
            ))
            if is_total:
                ax.plot([x, x + cw], [ry + rh, ry + rh],
                        color=_C["total_border"], linewidth=1.5, zorder=3)

            lines_      = wr[c]
            _, color    = _fmt_cell(row[c], col_numeric[c])
            if col_account[c] and _ACCT_RE.match(row[c].strip()):
                color = _C["account"]

            align = "right" if col_numeric[c] else "left"
            tx    = (x + cw - PAD_H) if col_numeric[c] else (x + PAD_H)
            n     = len(lines_)
            block_h    = n * LINE_H
            top_line_y = ry + (rh + block_h) / 2 - LINE_H / 2

            for ln_idx, ln in enumerate(lines_):
                ly = top_line_y - ln_idx * LINE_H
                ax.text(tx, ly, ln,
                        ha=align, va="center", fontsize=8.5,
                        fontweight="bold" if is_total else "normal",
                        color=color, fontfamily="DejaVu Sans", zorder=2)
            x += cw

    return fig


# ══════════════════════════════════════════════════════════════════════════════
# PNG export
# ══════════════════════════════════════════════════════════════════════════════

def render_table_png(
    rows: list[list[str]], dpi: int = 150, title: str | None = None
) -> bytes:
    if not rows:
        return b""
    fig = _build_figure(rows, dpi, title=_strip_emoji(title))
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=dpi,
                bbox_inches="tight", facecolor=_C["canvas"], pad_inches=0.10)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Excel export  (ported from accountant_bot/bot/table_render.py)
# ══════════════════════════════════════════════════════════════════════════════

def render_tables_xlsx(
    tables: list[list[list[str]]],
    titles: list[str | None] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="1E2D45")
    HEADER_FONT  = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
    TITLE_FILL   = PatternFill("solid", fgColor="2C3E50")
    TITLE_FONT   = Font(name="Calibri", size=12, bold=True,  color="FFFFFF")
    TOTAL_FILL   = PatternFill("solid", fgColor="E4EBF5")
    EVEN_FILL    = PatternFill("solid", fgColor="F2F5FA")
    ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")
    TOTAL_FONT   = Font(name="Calibri", size=11, bold=True,  color="1A2332")
    DATA_FONT    = Font(name="Calibri", size=11,              color="1A2332")
    NEG_FONT     = Font(name="Calibri", size=11,              color="C0392B")
    ACCT_FONT    = Font(name="Calibri", size=11,              color="2563A8")
    MUTED_FONT   = Font(name="Calibri", size=11,              color="9BA8B5")

    thin  = Side(style="thin",   color="DDE3EC")
    thick = Side(style="medium", color="1E2D45")
    CELL_BORDER  = Border(left=thin, right=thin, top=thin,  bottom=thin)
    TOTAL_BORDER = Border(left=thin, right=thin, top=thick, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    for t_idx, rows in enumerate(tables):
        if not rows:
            continue
        title = (titles[t_idx] if titles and t_idx < len(titles) else None)
        n_cols  = max(len(r) for r in rows)
        norm    = [(r + [""] * n_cols)[:n_cols] for r in rows]
        header  = norm[0]
        d_rows  = norm[1:] if len(norm) > 1 else [[""] * n_cols]

        col_numeric, col_account = _detect_col_types(d_rows, n_cols)
        total_set = {i for i, r in enumerate(d_rows) if _is_total_row(r)}

        sheet_name = f"Таблица {t_idx + 1}" if len(tables) > 1 else "Данные"
        ws = wb.create_sheet(title=sheet_name)

        start_row = 1
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            cell = ws.cell(row=1, column=1, value=title)
            cell.fill      = TITLE_FILL
            cell.font      = TITLE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 24
            start_row = 2

        for c, val in enumerate(header, start=1):
            cell = ws.cell(row=start_row, column=c, value=val)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="right" if col_numeric[c - 1] else "left",
                vertical="center",
            )
        ws.row_dimensions[start_row].height = 22

        for i, row in enumerate(d_rows):
            xl_row     = start_row + 1 + i
            is_total   = i in total_set
            fill       = TOTAL_FILL if is_total else (ODD_FILL if i % 2 == 0 else EVEN_FILL)
            border     = TOTAL_BORDER if is_total else CELL_BORDER

            for c, val in enumerate(row, start=1):
                is_num = col_numeric[c - 1]
                is_acc = col_account[c - 1]
                display, color_key = _fmt_cell(val, is_num)

                xl_val: object = display
                if is_num and display not in ("—", ""):
                    raw = display.replace(" ", "").replace(" ", "").replace(",", ".")
                    raw = raw.strip("()")
                    try:
                        xl_val = -float(raw) if display.startswith("(") else float(raw)
                    except ValueError:
                        xl_val = display

                cell        = ws.cell(row=xl_row, column=c, value=xl_val)
                cell.fill   = fill
                cell.border = border

                if display == "—":
                    cell.font = MUTED_FONT
                elif is_total:
                    cell.font = TOTAL_FONT
                elif color_key == _C["negative"]:
                    cell.font = NEG_FONT
                elif is_acc and _ACCT_RE.match(val.strip()):
                    cell.font = ACCT_FONT
                else:
                    cell.font = DATA_FONT

                if is_num and isinstance(xl_val, float):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(
                        horizontal="right" if is_num else "left",
                        vertical="center",
                    )
            ws.row_dimensions[xl_row].height = 18

        for c in range(1, n_cols + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=c).value or ""))
                for r in range(1, len(norm) + start_row)
            )
            ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 4, 10), 45)

        ws.freeze_panes = f"A{start_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# reportlab font registration
# ══════════════════════════════════════════════════════════════════════════════

def _register_dejavu() -> str:
    """Register DejaVu Sans (bundled with matplotlib) in reportlab. Returns font name."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from matplotlib.font_manager import findfont, FontProperties

    variants = {
        "DejaVuSans":            FontProperties(family="DejaVu Sans", weight="normal", style="normal"),
        "DejaVuSans-Bold":       FontProperties(family="DejaVu Sans", weight="bold",   style="normal"),
        "DejaVuSans-Oblique":    FontProperties(family="DejaVu Sans", weight="normal", style="italic"),
        "DejaVuSans-BoldOblique":FontProperties(family="DejaVu Sans", weight="bold",   style="italic"),
    }
    registered = []
    for name, fp in variants.items():
        try:
            path = findfont(fp)
            pdfmetrics.registerFont(TTFont(name, path))
            registered.append(name)
        except Exception:
            pass

    if "DejaVuSans" in registered:
        pdfmetrics.registerFontFamily(
            "DejaVuSans",
            normal="DejaVuSans",
            bold="DejaVuSans-Bold"       if "DejaVuSans-Bold"       in registered else "DejaVuSans",
            italic="DejaVuSans-Oblique"  if "DejaVuSans-Oblique"    in registered else "DejaVuSans",
            boldItalic="DejaVuSans-BoldOblique" if "DejaVuSans-BoldOblique" in registered else "DejaVuSans",
        )
        return "DejaVuSans"
    return "Helvetica"


# ══════════════════════════════════════════════════════════════════════════════
# Markdown → reportlab XML  (mini shim for PDF prose)
# ══════════════════════════════════════════════════════════════════════════════

_RL_BOLD   = re.compile(r"\*\*(.+?)\*\*")
_RL_CODE   = re.compile(r"`([^`]+)`")
_RL_CODE3  = re.compile(r"```[\w]*\n?([\s\S]*?)```")
_RL_HEADING = re.compile(r"^(#{1,3})\s+(.*)")
_RL_BULLET  = re.compile(r"^[\•\-\*]\s+(.+)")
_RL_NUMBERED = re.compile(r"^\d+\.\s+(.+)")


def _para_xml(text: str, font: str) -> str:
    """Escape HTML special chars and apply markdown→reportlab XML."""
    t = _html_mod.escape(text)
    t = _RL_BOLD.sub(r"<b>\1</b>", t)
    t = _RL_CODE.sub(r'<font name="Courier">\1</font>', t)
    return t


def _md_to_flowables(text: str, font: str):
    """Convert a markdown text block to a list of reportlab Platypus flowables."""
    from reportlab.platypus import Paragraph, Spacer, Preformatted
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_LEFT

    normal_style = ParagraphStyle(
        "normal_ru", fontName=font, fontSize=10, leading=14,
        spaceAfter=4, spaceBefore=2,
    )
    h1_style = ParagraphStyle(
        "h1_ru", fontName=font, fontSize=15, leading=20,
        spaceAfter=6, spaceBefore=10, textColor="#1E2D45",
    )
    h2_style = ParagraphStyle(
        "h2_ru", fontName=font, fontSize=13, leading=18,
        spaceAfter=5, spaceBefore=8, textColor="#2C3E50",
    )
    h3_style = ParagraphStyle(
        "h3_ru", fontName=font, fontSize=11, leading=15,
        spaceAfter=4, spaceBefore=6, textColor="#2C3E50",
    )
    bullet_style = ParagraphStyle(
        "bullet_ru", fontName=font, fontSize=10, leading=14,
        spaceAfter=2, leftIndent=16,
        bulletIndent=4,
    )
    code_style = ParagraphStyle(
        "code_ru", fontName="Courier", fontSize=9, leading=13,
        spaceAfter=4, backColor="#F2F5FA",
    )

    flowables = []

    # Handle triple-backtick code blocks first
    remaining = text
    code_parts = _RL_CODE3.split(remaining)
    # split gives [before, code1, after, code2, after2, ...]
    segments_txt = []
    for idx, part in enumerate(code_parts):
        if idx % 2 == 0:
            segments_txt.append(("text", part))
        else:
            segments_txt.append(("code", part))

    for kind, block in segments_txt:
        if kind == "code":
            for line in block.strip().split("\n"):
                flowables.append(Preformatted(line, code_style))
            flowables.append(Spacer(1, 4))
            continue

        for line in block.split("\n"):
            line_stripped = line.strip()

            if not line_stripped:
                flowables.append(Spacer(1, 6))
                continue

            # Heading?
            hm = _RL_HEADING.match(line_stripped)
            if hm:
                level = len(hm.group(1))
                content = _para_xml(hm.group(2), font)
                style = [h1_style, h2_style, h3_style][min(level - 1, 2)]
                flowables.append(Paragraph(content, style))
                continue

            # Bullet?
            bm = _RL_BULLET.match(line_stripped) or _RL_NUMBERED.match(line_stripped)
            if bm:
                content = _para_xml(bm.group(1), font)
                flowables.append(Paragraph(f"• {content}", bullet_style))
                continue

            # Normal paragraph
            content = _para_xml(line_stripped, font)
            flowables.append(Paragraph(content, normal_style))

    return flowables


# ══════════════════════════════════════════════════════════════════════════════
# Rendering modes
# ══════════════════════════════════════════════════════════════════════════════

def mode_segments(answer: str) -> dict:
    """
    Returns ordered segments list where tables are rendered to PNG.
    """
    ordered = split_ordered(answer)
    result = []
    for seg in ordered:
        if seg["type"] == "text":
            result.append({"type": "text", "md": seg["md"]})
        else:
            png_bytes = render_table_png(seg["rows"], dpi=150, title=_strip_emoji(seg.get("title")))
            b64 = base64.b64encode(png_bytes).decode("ascii")
            result.append({"type": "png", "b64": b64})
    return {"segments": result}


def mode_xlsx(answer: str) -> dict:
    """Extract all wide tables and return as base64 xlsx workbook."""
    ordered = split_ordered(answer)
    tables = [seg["rows"] for seg in ordered if seg["type"] == "table"]
    titles = [seg.get("title") for seg in ordered if seg["type"] == "table"]
    if not tables:
        return {"b64": "", "empty": True}
    xlsx_bytes = render_tables_xlsx(tables, titles)
    return {"b64": base64.b64encode(xlsx_bytes).decode("ascii")}


def mode_pdf(answer: str) -> dict:
    """
    Build a full-answer PDF: prose text + table images interleaved.
    Uses reportlab Platypus with DejaVu Sans for Cyrillic support.
    """
    from reportlab.platypus import SimpleDocTemplate, Image as RLImage, Spacer, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm

    try:
        font = _register_dejavu()
    except Exception:
        font = "Helvetica"

    ordered = split_ordered(answer)
    flowables = []
    W, H = A4

    for seg in ordered:
        if seg["type"] == "text":
            flowables.extend(_md_to_flowables(seg["md"], font))

        elif seg["type"] == "table":
            png_bytes = render_table_png(seg["rows"], dpi=150, title=_strip_emoji(seg.get("title")))
            if not png_bytes:
                continue
            img_buf = BytesIO(png_bytes)

            # Determine rendered figure dimensions and scale to fit page width
            try:
                from PIL import Image as PILImage
                pil = PILImage.open(BytesIO(png_bytes))
                img_w_px, img_h_px = pil.size
            except ImportError:
                # Fall back: assume 150 dpi
                import struct
                # Read PNG IHDR
                img_w_px = struct.unpack(">I", png_bytes[16:20])[0]
                img_h_px = struct.unpack(">I", png_bytes[20:24])[0]

            # Convert pixels at 150 dpi → reportlab points (1 inch = 72 pt).
            # Scale down only if wider than the printable area; never upscale.
            rl_w_full = (img_w_px / 150.0) * 72  # pt
            rl_h_full = (img_h_px / 150.0) * 72  # pt
            max_w = W - 3 * cm  # printable width (A4 minus 1.5 cm each side)
            scale = min(1.0, max_w / rl_w_full)
            rl_w  = rl_w_full * scale
            rl_h  = rl_h_full * scale

            flowables.append(Spacer(1, 6))
            flowables.append(RLImage(BytesIO(png_bytes), width=rl_w, height=rl_h))
            flowables.append(Spacer(1, 6))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm,  bottomMargin=1.5 * cm,
    )
    doc.build(flowables)
    buf.seek(0)
    return {"b64": base64.b64encode(buf.read()).decode("ascii")}


# ══════════════════════════════════════════════════════════════════════════════
# CLI entry point
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    if len(sys.argv) < 2:
        sys.stdout.buffer.write(json.dumps({"error": "Usage: render_cli.py <segments|xlsx|pdf>"}).encode("utf-8"))
        sys.exit(1)

    mode = sys.argv[1].strip()

    # Read stdin as raw UTF-8 bytes — avoids Windows cp1251/cp1252 console default
    # corrupting Cyrillic text in the JSON payload.
    payload = json.loads(sys.stdin.buffer.read().decode("utf-8"))
    answer  = payload.get("answer", "")

    try:
        if mode == "segments":
            result = mode_segments(answer)
        elif mode == "xlsx":
            result = mode_xlsx(answer)
        elif mode == "pdf":
            result = mode_pdf(answer)
        else:
            result = {"error": f"Unknown mode: {mode}"}
    except Exception as exc:
        import traceback
        print(traceback.format_exc(), file=sys.stderr)
        result = {"error": str(exc)}

    # Write output as raw UTF-8 bytes for the same reason as stdin above.
    sys.stdout.buffer.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
    sys.stdout.buffer.flush()


if __name__ == "__main__":
    main()
